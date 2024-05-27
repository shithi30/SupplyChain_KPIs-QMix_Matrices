#!/usr/bin/env python
# coding: utf-8

# In[1]:


# import
import pandas as pd
import duckdb
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
import win32com.client
from win32com.client import Dispatch
from datetime import datetime
import time


# In[2]:


# read target/Stock Alloc
def read_tgt_file():
    file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/RPL Inputs/" + "February'24 Town x SKU Stock Allocation_National.xlsx"
    sheet_name = "Town x SKU x Case x TGT "
    df = pd.read_excel(open(file, "rb"), sheet_name=sheet_name, header=2, index_col=None)
    tgt_df = df[['CATEGORY', 'TOWN NAME', 'SKU NAME', 'TOWN x SKU TGT - TP Cr.']]
    tgt_df.columns = ['category', 'town', 'basepack', 'tgt_cr']
    tgt_df = duckdb.query('''select upper(category) category, upper(town) town, upper(basepack) basepack, tgt_cr from tgt_df''').df()
    return tgt_df


# In[3]:


# read RPL
def fetch_read_rpl(rec_date_from, rec_date_to): 
    
    # inputs
    subject_pattern = 'Replenishment Report'
    atch_pattern = 'Replenishment Repot'

    # output folder
    output_dir = Path.cwd() / 'RPL Inputs'
    output_dir.mkdir(parents=True, exist_ok=True)

    # output files
    filenames = []

    # outlook inbox
    outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.Folders.Item(1).Folders['Kader Bhai']

    # emails
    messages = inbox.Items
    for message in reversed(messages): 

        # time
        try: rec_date = str(message.SentOn)[0:10]
        except: continue
        if rec_date < rec_date_from or rec_date > rec_date_to: continue

        # subject
        subject = message.Subject
        if subject_pattern.lower() not in subject.lower(): continue

        # attachments
        attachments = message.Attachments
        for attachment in attachments:
            filename = rec_date + '_' + attachment.FileName
            if atch_pattern.lower() in filename.lower(): 
                filenames.append(filename)
                attachment.SaveAsFile(output_dir / filename)

    # read
    rpl_df = pd.DataFrame()
    for f in filenames:
        print("Reading: " + f)
        file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/RPL Inputs/" + f
        df = pd.read_excel(open(file, "rb"), sheet_name="Replenishment UBL_UCL", header=0, index_col=None)
        df = df[['Date', 'Town', 'Basepack', 'Proposed qty', 'Norm qty']]
        df.columns = ['rpl_date', 'town', 'basepack', 'proposed_qty', 'norm_qty']
        df = duckdb.query('''select strptime(rpl_date, '%d %b %Y') rpl_date, upper(town) town, upper(basepack) basepack, proposed_qty, norm_qty from df''').df()
        rpl_df = rpl_df.append(df)
    return rpl_df


# In[4]:


# prepare report
def prepare_report(rec_date_from, rec_date_to, rpl_df, sheet_name):
    
    # prepare RPL
    rpl_df = rpl_df.fillna(-0.01)
    qry = '''
    select 
        case when basepack is not null then basepack else 'unspecified' end basepack, 
        case when category is not null then category else 'unspecified' end category, 
        case when town is not null then town else 'unspecified' end town, 
        proposed_qty, norm_qty, qmix, 
        case when bp_tgt_cr*1.00/tot_tgt_cr is null then 0 else bp_tgt_cr*1.00/tot_tgt_cr end bp_bc, 
        case when town_tgt_cr*1.00/tot_tgt_cr is null then 0 else town_tgt_cr*1.00/tot_tgt_cr end town_bc
    from 
        (select basepack, town, sum(proposed_qty) proposed_qty, sum(norm_qty) norm_qty, 1-sum(proposed_qty)*1.00/sum(norm_qty) qmix
        from rpl_df
        group by 1, 2
        ) tbl0

        full join

        (select basepack, sum(tgt_cr) bp_tgt_cr
        from tgt_df
        group by 1
        ) tbl2 using(basepack)

        full join

        (select town, sum(tgt_cr) town_tgt_cr
        from tgt_df
        group by 1
        ) tbl3 using(town)

        left join 

        (select distinct basepack, category
        from tgt_df
        ) tbl4 using(basepack), 

        (select sum(tgt_cr) tot_tgt_cr
        from tgt_df
        ) tbl5
    order by bp_bc desc
    '''
    rpl_df_contrib = duckdb.query(qry).df()

    # order columns
    qry = '''
    select 
        town, 
        max(town_bc) town_bc, 
        sum(max(town_bc)) over(order by max(town_bc) desc) town_bc_cum,
        1-sum(proposed_qty)*1.00/sum(norm_qty) town_qmix
    from rpl_df_contrib
    group by 1
    order by town_bc desc
    '''
    ord_df = duckdb.query(qry).df()
    town_ord = ord_df['town'].tolist()

    # town BC
    town_bc_cum_ord = ['town_bc_cum'] + ord_df['town_bc_cum'].tolist()
    town_bc_cum_ord = pd.DataFrame([town_bc_cum_ord])
    town_bc_ord = ['town_bc'] + ord_df['town_bc'].tolist()
    town_bc_ord = pd.DataFrame([town_bc_ord])
    town_qmix_ord = ['town_qmix'] + ord_df['town_qmix'].tolist()
    town_qmix_ord = pd.DataFrame([town_qmix_ord])

    # basepack BC
    qry = '''
    select 
        category, 
        basepack, 
        max(bp_bc) basepack_bc, 
        sum(max(bp_bc)) over(order by max(bp_bc) desc) basepack_bc_cum,
        1-sum(proposed_qty)*1.00/sum(norm_qty) basepack_qmix
    from rpl_df_contrib
    group by 1, 2
    order by basepack_bc desc
    '''
    bp_bc_df = duckdb.query(qry).df()[['basepack_bc_cum', 'basepack_bc', 'basepack_qmix', 'category']]
    
    # national
    qry = '''
    select * 
    from 
        (select 1-sum(proposed_qty)*1.00/sum(norm_qty) national_qmix
        from rpl_df_contrib
        ) tbl1, 
        (select 1-sum(proposed_qty)*1.00/sum(norm_qty) national_qmix_excluding_mtsmtwt
        from rpl_df_contrib
        where town not in('WATER-DHAKA', 'MODERN TRADE', 'SMT & SHOPPING COMPL', 'OOH DISTRIBUTOR DHAK')
        ) tbl2
    '''
    national_df = duckdb.query(qry).df()

    # pivot
    rpl_df_piv = pd.pivot_table(rpl_df_contrib.fillna(-0.01), values='qmix', index=['basepack'], columns='town', sort=False).fillna(-0.01)
    rpl_df_piv = rpl_df_piv[town_ord]
    
    # path
    path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/QMix Matrices/QMix_2by2_Matrix_' + rec_date_from[0:7] + '.xlsx'

    # if exists
    if_exists = 1
    try: book = load_workbook(path)
    except: if_exists = 0

    # writer
    writer = pd.ExcelWriter(path, engine = 'openpyxl')

    # create/retrieve/remove sheet(s)
    if if_exists == 1: 
        writer.book = book
        if sheet_name in book.sheetnames:
            if rec_date_from != rec_date_to:
                try: del writer.book[sheet_name] # sheet exists 
                except: pass                     # sheet does not exist
            else: 
                writer.close()
                return
         
    # write
    national_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=3, index=False)
    town_bc_cum_ord.to_excel(writer, sheet_name=sheet_name, startrow=2, startcol=4, header=False, index=False)
    town_bc_ord.to_excel(writer, sheet_name=sheet_name, startrow=3, startcol=4, header=False, index=False)
    town_qmix_ord.to_excel(writer, sheet_name=sheet_name, startrow=4, startcol=4, header=False, index=False)
    rpl_df_piv.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=4)
    bp_bc_df.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=0, index=False)

    # adjust
    worksheet = writer.sheets[sheet_name]
    for column_cells in worksheet.columns:
        length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)        
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    writer.close()
    
    # format
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    
    # percent
    column_letters = [col.column_letter for col in worksheet[1]]
    for col in column_letters: 
        for cel in worksheet[(col)]: 
            cel.number_format = "0.00%"
    
    # color
    font = Font(bold = True, color = 'EE1111')
    dxf = DifferentialStyle(font = font)
    rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, national_df['national_qmix'].tolist()[0]], dxf = dxf)
    worksheet.conditional_formatting.add('C5:GZ300', rule)
    
    # freeze
    worksheet.freeze_panes = worksheet['F7']
    
    # save
    workbook.save(path)

    # tabs in sheet
    print("Worksheets in workbook: ")
    print(pd.ExcelFile(path).sheet_names)
    print()


# In[5]:


# record
start_time = time.time()


# In[6]:


# month, dates, target
tgt_df = read_tgt_file()

current_month = datetime.today().strftime('%Y-%m-%d')[0:7]
qry = '''
select (concat(left(current_date, 7), '-01')::date + generate_series::int)::text qmix_date
from (select * from generate_series(0, 100)) tbl1 
where generate_series < date_part('day', current_date)
'''
qmix_dates = duckdb.query(qry).df()['qmix_date'].tolist()

# current_month = '2024-01'
# qmix_dates = ['2024-01-01', '2024-01-02', '2024-01-03', '2024-01-04', '2024-01-05', '2024-01-06', '2024-01-07', '2024-01-08', '2024-01-09', '2024-01-10', '2024-01-11', '2024-01-12', '2024-01-13', '2024-01-14', '2024-01-15', '2024-01-16', '2024-01-17', '2024-01-18', '2024-01-19', '2024-01-20', '2024-01-21', '2024-01-22', '2024-01-23', '2024-01-24', '2024-01-25', '2024-01-26', '2024-01-27', '2024-01-28', '2024-01-29', '2024-01-30', '2024-01-31']

print(qmix_dates)


# In[7]:


# RPL
rec_date_from = qmix_dates[0]
rec_date_to = qmix_dates[len(qmix_dates)-1]
rpl_df = fetch_read_rpl(rec_date_from, rec_date_to)


# In[8]:


# # RPL
# rec_date_from = qmix_dates[0]
# rec_date_to = qmix_dates[len(qmix_dates)-1]
# filenames = [
#     'Replenishment Repot_01 Jan 2024.xlsx',
#     'Replenishment Repot_02 Jan 2024.xlsx',
#     'Replenishment Repot_03 Jan 2024.xlsx',
    
#     'Replenishment Repot_04 Jan 2024.xlsx',
#     'Replenishment Repot_05 Jan 2024.xlsx',
#     'Replenishment Repot_06 Jan 2024.xlsx',
    
#     'Replenishment Repot_07 Jan 2024.xlsx',
#     'Replenishment Repot_08 Jan 2024.xlsx',
#     'Replenishment Repot_08 Jan 2024.xlsx',
    
#     'Replenishment Repot_10 Jan 2024.xlsx',
#     'Replenishment Repot_11 Jan 2024.xlsx',
#     'Replenishment Repot_12 Jan 2024.xlsx',
    
#     'Replenishment Repot_13 Jan 2024.xlsx',
#     'Replenishment Repot_14 Jan 2024.xlsx',
#     'Replenishment Repot_15 Jan 2024.xlsx',
    
#     'Replenishment Repot_16 Jan 2024.xlsx',
#     'Replenishment Repot_17 Jan 2024.xlsx',
#     'Replenishment Repot_18 Jan 2024.xlsx',
    
#     'Replenishment Repot_19 Jan 2024.xlsx',
#     'Replenishment Repot_20 Jan 2024.xlsx',
#     'Replenishment Repot_21 Jan 2024.xlsx',
    
#     'Replenishment Repot_22 Jan 2024.xlsx',
#     'Replenishment Repot_23 Jan 2024.xlsx',
#     'Replenishment Repot_24 Jan 2024.xlsx',
    
#     'Replenishment Repot_25 Jan 2024.xlsx',
#     'Replenishment Repot_26 Jan 2024.xlsx',
#     'Replenishment Repot_27 Jan 2024.xlsx',
    
#     'Replenishment Repot_28 Jan 2024.xlsx',
#     'Replenishment Repot_29 Jan 2024.xlsx',
#     'Replenishment Repot_30 Jan 2024.xlsx',
    
#     'Replenishment Repot_31 Jan 2024.xlsx'
# ]
# rpl_df = pd.DataFrame()
# for f in filenames:
#     file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/RPL Inputs/" + f
#     print(1)
#     try: df = pd.read_excel(open(file, "rb"), sheet_name="Replenishment UBL_UCL", header=0, index_col=None)
#     except: continue
#     print("Reading: " + f)
#     df = df[['Date', 'Town', 'Basepack', 'Proposed qty', 'Norm qty']]
#     df.columns = ['rpl_date', 'town', 'basepack', 'proposed_qty', 'norm_qty']
#     df = duckdb.query('''select strptime(rpl_date, '%d %b %Y') rpl_date, upper(town) town, upper(basepack) basepack, proposed_qty, norm_qty from df''').df()
#     rpl_df = rpl_df.append(df)


# In[9]:


# daily
for qmix_date in qmix_dates:
    day_df = duckdb.query("select * from rpl_df where rpl_date='" + qmix_date + "'").df()
    if day_df.shape[0] == 0: continue
    prepare_report(qmix_date, qmix_date, day_df, qmix_date)


# In[10]:


# MTD
prepare_report(rec_date_from, rec_date_to, rpl_df, current_month + "-MTD")


# In[11]:


# TDP-01
rec_date_from = current_month + "-01"
rec_date_to = current_month + "-10"
tdp_df = duckdb.query("select * from rpl_df where rpl_date>='" + rec_date_from + "' and rpl_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP1")

# TDP-02
rec_date_from = current_month + "-11"
rec_date_to = current_month + "-20"
tdp_df = duckdb.query("select * from rpl_df where rpl_date>='" + rec_date_from + "' and rpl_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP2")

# # TDP-03
# rec_date_from = current_month + "-21"
# rec_date_to = qmix_dates[len(qmix_dates)-1]
# tdp_df = duckdb.query("select * from rpl_df where rpl_date>='" + rec_date_from + "' and rpl_date<='" + rec_date_to + "'").df()
# prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP3")


# In[12]:


## trend ##

# path
path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/QMix Matrices/QMix_2by2_Matrix_' + current_month + '.xlsx'

# sheets
sheets = pd.ExcelFile(path).sheet_names

# trend data
qry = '''
select left(rpl_date::text, 10) qmix_date, basepack, town, 1-sum(proposed_qty)*1.00/sum(norm_qty) qmix
from rpl_df
group by 1, 2, 3
union all 
select left(rpl_date::text, 10) qmix_date, 'overall' basepack, town, 1-sum(proposed_qty)*1.00/sum(norm_qty) qmix
from rpl_df
group by 1, 2, 3
union all 
select left(rpl_date::text, 10) qmix_date, basepack, 'overall' town, 1-sum(proposed_qty)*1.00/sum(norm_qty) qmix
from rpl_df
group by 1, 2, 3
'''
trend_df = duckdb.query(qry).df()
trend_df_piv = pd.pivot_table(trend_df.fillna(-0.01), values='qmix', index=['town', 'basepack'], columns='qmix_date', sort=False).fillna(-0.01)

# benchmark
qry = '''
select 1-sum(proposed_qty)*1.00/sum(norm_qty) benchmark_qmix
from rpl_df
'''
benchmark = duckdb.query(qry).df()['benchmark_qmix'].tolist()[0] 

# load
book = load_workbook(path)

# writer
writer = pd.ExcelWriter(path, engine = 'openpyxl')

# create, retrieve, remove sheets
sheet_name = current_month + '-trends'
writer.book = book
if sheet_name in sheets: del writer.book[sheet_name]

# write
trend_df_piv.reset_index().to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index = False)

# adjust
worksheet = writer.sheets[sheet_name]
it = 1
for column_cells in worksheet.columns:
    if it in [1, 2]: length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)  
    else: length = len(str(column_cells[0].value))
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    it = it + 1
writer.close()

# format
workbook = load_workbook(path)
worksheet = workbook[sheet_name]

# freeze
worksheet.freeze_panes = worksheet['C2']

# filter
worksheet.auto_filter.ref = worksheet.dimensions

# percent
column_letters = [col.column_letter for col in worksheet[1]]
for col in column_letters: 
    for cel in worksheet[(col)]: 
        cel.number_format = "0.00%" 
        
# color
font = Font(bold = True, color = 'EE1111')
dxf = DifferentialStyle(font = font)
rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, benchmark], dxf = dxf)
worksheet.conditional_formatting.add('C5:AH50000', rule)
    
# save
workbook.save(path)

# tabs in sheet
print("Worksheets in workbook: ")
print(pd.ExcelFile(path).sheet_names)


# In[13]:


# report
elapsed_time = str(round((time.time() - start_time) / 60.00, 2))
print("Elapsed time to run script (mins): " + elapsed_time)


# In[ ]:




