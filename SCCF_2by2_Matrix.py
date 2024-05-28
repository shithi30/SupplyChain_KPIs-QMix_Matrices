#!/usr/bin/env python
# coding: utf-8

# In[1]:


# import
import pandas as pd
import duckdb
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
import win32com.client
from win32com.client import Dispatch
from datetime import datetime
import time


# In[2]:


# read target/Stock Alloc
def read_tgt_file():
    file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/RPL Inputs/" + "February'24 Town x SKU Stock Allocation_National.xlsx"
    sheet_name = "Town x SKU x Case x TGT "
    df = pd.read_excel(open(file, "rb"), sheet_name=sheet_name, header=2, index_col=None)
    tgt_df = df[['CATEGORY', 'TOWN NAME', 'SKU NAME', 'TOWN x SKU TGT - TP Cr.']]
    tgt_df.columns = ['category', 'town', 'basepack', 'tgt_cr']
    tgt_df = duckdb.query('''select upper(category) category, upper(town) town, upper(basepack) basepack, tgt_cr from tgt_df''').df()
    return tgt_df


# In[3]:


# read SCCF
def fetch_read_sccf(rec_date_from, rec_date_to): 
    
    # inputs
    subject_pattern = 'Secondary CCFOT'

    # output folder
    output_dir = Path.cwd() / 'SCCF Inputs'
    output_dir.mkdir(parents=True, exist_ok=True)

    # output files
    filenames = set()

    # outlook inbox
    outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.Folders.Item(1).Folders['Kader Bhai']

    # emails
    messages = inbox.Items
    for message in reversed(messages): 

        # subject
        subject = message.Subject
        if subject_pattern.lower() not in subject.lower(): continue

        # attachments
        attachments = message.Attachments
        for attachment in attachments:
            filename = attachment.FileName
            try: file_date = str(datetime.strptime(" ".join(filename.split()[3:6])[0:-9], '%d %b %Y'))[0:10]
            except: file_date = str(datetime.strptime(" ".join(filename.split()[3:6])[0:-9], '%d %B %Y'))[0:10]
            if file_date >= rec_date_from and file_date <= rec_date_to:
                filenames.add(filename)
                attachment.SaveAsFile(output_dir / filename)
                
    # read
    sccf_df = pd.DataFrame()
    for f in filenames:
        print("Reading: " + f)
        file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/SCCF Inputs/" + f
        df = pd.read_excel(open(file, "rb"), sheet_name="Sheet1", header=1, index_col=None)
        df = df[['Local Sales Region 4', 'Pack Size', 'CS', 'CS.1']]
        df.columns = ['town', 'basepack', 'ord_qty', 'inv_qty']
        df = duckdb.query('''select upper(town) town, upper(basepack) basepack, ord_qty, inv_qty from df''').df()
        try: df['sccf_date'] = str(datetime.strptime(" ".join(f.split()[3:6])[0:-9], '%d %B %Y'))[0:10]
        except: df['sccf_date'] = str(datetime.strptime(" ".join(f.split()[3:6])[0:-9], '%d %b %Y'))[0:10]
        sccf_df = sccf_df.append(df)
    return sccf_df


# In[4]:


# prepare report
def prepare_report(rec_date_from, rec_date_to, sccf_df, sheet_name):
    
    # prepare SCCF
    sccf_df = sccf_df.fillna(-0.01)
    qry = '''
    select 
        case when basepack is not null then basepack else 'unspecified' end basepack, 
        case when category is not null then category else 'unspecified' end category, 
        case when town is not null then town else 'unspecified' end town, 
        ord_qty, inv_qty, sccf,
        case when bp_tgt_cr*1.00/tot_tgt_cr is null then 0 else bp_tgt_cr*1.00/tot_tgt_cr end bp_bc, 
        case when town_tgt_cr*1.00/tot_tgt_cr is null then 0 else town_tgt_cr*1.00/tot_tgt_cr end town_bc
    from 
        (select basepack, town, sum(ord_qty) ord_qty, sum(inv_qty) inv_qty, sum(inv_qty)*1.00/sum(ord_qty) sccf
        from sccf_df
        group by 1, 2
        ) tbl0

        full join

        (select basepack, sum(tgt_cr) bp_tgt_cr
        from tgt_df
        group by 1
        ) tbl2 using(basepack)

        full join

        (select town, sum(tgt_cr) town_tgt_cr
        from tgt_df
        group by 1
        ) tbl3 using(town)

        left join 

        (select distinct basepack, category
        from tgt_df
        ) tbl4 using(basepack), 

        (select sum(tgt_cr) tot_tgt_cr
        from tgt_df
        ) tbl5
    order by bp_bc desc
    '''
    sccf_df_contrib = duckdb.query(qry).df()

    # order columns
    qry = '''
    select 
        town, 
        max(town_bc) town_bc, 
        sum(max(town_bc)) over(order by max(town_bc) desc) town_bc_cum,
        sum(inv_qty)*1.00/sum(ord_qty) town_sccf
    from sccf_df_contrib
    group by 1
    order by town_bc desc
    '''
    ord_df = duckdb.query(qry).df()
    town_ord = ord_df['town'].tolist()

    # town BC
    town_bc_cum_ord = ['town_bc_cum'] + ord_df['town_bc_cum'].tolist()
    town_bc_cum_ord = pd.DataFrame([town_bc_cum_ord])
    town_bc_ord = ['town_bc'] + ord_df['town_bc'].tolist()
    town_bc_ord = pd.DataFrame([town_bc_ord])
    town_sccf_ord = ['town_sccf'] + ord_df['town_sccf'].tolist()
    town_sccf_ord = pd.DataFrame([town_sccf_ord])

    # basepack BC
    qry = '''
    select 
        category, 
        basepack, 
        max(bp_bc) basepack_bc, 
        sum(max(bp_bc)) over(order by max(bp_bc) desc) basepack_bc_cum,
        sum(inv_qty)*1.00/sum(ord_qty) basepack_sccf
    from sccf_df_contrib
    group by 1, 2
    order by basepack_bc desc
    '''
    bp_bc_df = duckdb.query(qry).df()[['basepack_bc_cum', 'basepack_bc', 'basepack_sccf', 'category']]
    
    # national
    qry = '''
    select * 
    from 
        (select sum(inv_qty)*1.00/sum(ord_qty) national_sccf
        from sccf_df_contrib
        ) tbl1, 
        (select sum(inv_qty)*1.00/sum(ord_qty) national_sccf_excluding_mtsmtwt
        from sccf_df_contrib
        where town not in('WATER-DHAKA', 'MODERN TRADE', 'SMT & SHOPPING COMPL', 'OOH DISTRIBUTOR DHAK')
        ) tbl2
    '''
    national_df = duckdb.query(qry).df()

    # pivot
    sccf_df_piv = pd.pivot_table(sccf_df_contrib.fillna(-0.01), values='sccf', index=['basepack'], columns='town', sort=False).fillna(-0.01)
    sccf_df_piv = sccf_df_piv[town_ord]
    
    # path
    path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/SCCF Matrices/SCCF_2by2_Matrix_' + rec_date_from[0:7] + '.xlsx'

    # if exists
    if_exists = 1
    try: book = load_workbook(path)
    except: if_exists = 0

    # writer
    writer = pd.ExcelWriter(path, engine = 'openpyxl')

    # create/retrieve/remove sheet(s)
    if if_exists == 1: 
        writer.book = book
        if sheet_name in book.sheetnames:
            if rec_date_from != rec_date_to:
                try: del writer.book[sheet_name] # sheet exists 
                except: pass                     # sheet does not exist
            else: 
                writer.close()
                return
         
    # write
    national_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=3, index=False)
    town_bc_cum_ord.to_excel(writer, sheet_name=sheet_name, startrow=2, startcol=4, header=False, index=False)
    town_bc_ord.to_excel(writer, sheet_name=sheet_name, startrow=3, startcol=4, header=False, index=False)
    town_sccf_ord.to_excel(writer, sheet_name=sheet_name, startrow=4, startcol=4, header=False, index=False)
    sccf_df_piv.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=4)
    bp_bc_df.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=0, index=False)

    # adjust
    worksheet = writer.sheets[sheet_name]
    for column_cells in worksheet.columns:
        length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)        
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    writer.close()
    
    # format
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    
    # percent
    column_letters = [col.column_letter for col in worksheet[1]]
    for col in column_letters: 
        for cel in worksheet[(col)]: 
            cel.number_format = "0.00%"
    
    # color
    font = Font(bold = True, color = 'EE1111')
    dxf = DifferentialStyle(font = font)
    rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, national_df['national_sccf'].tolist()[0]], dxf = dxf)
    worksheet.conditional_formatting.add('C5:GZ300', rule)
    
    # freeze
    worksheet.freeze_panes = worksheet['F7']
    
    # save
    workbook.save(path)
    workbook.close()

    # tabs in sheet
    print("Worksheets in workbook: ")
    print(pd.ExcelFile(path).sheet_names)
    print()


# In[5]:


# record
start_time = time.time()


# In[6]:


# month, dates, target
tgt_df = read_tgt_file()

current_month = datetime.today().strftime('%Y-%m-%d')[0:7]
qry = '''
select (concat(left(current_date-2, 7), '-01')::date + generate_series::int)::text sccf_date
from (select * from generate_series(0, 100)) tbl1 
where generate_series < date_part('day', current_date)
'''
sccf_dates = duckdb.query(qry).df()['sccf_date'].tolist()
print(sccf_dates)


# In[7]:


# SCCF
rec_date_from = sccf_dates[0]
rec_date_to = sccf_dates[len(sccf_dates)-1]
sccf_df = fetch_read_sccf(rec_date_from, rec_date_to)


# In[8]:


# daily
for sccf_date in sccf_dates:
    day_df = duckdb.query("select * from sccf_df where sccf_date='" + sccf_date + "'").df()
    if day_df.shape[0] == 0: continue
    prepare_report(sccf_date, sccf_date, day_df, sccf_date)


# In[9]:


# MTD
prepare_report(rec_date_from, rec_date_to, sccf_df, current_month + "-MTD")


# In[15]:


# TDP-01
rec_date_from = current_month + "-01"
rec_date_to = current_month + "-10"
tdp_df = duckdb.query("select * from sccf_df where sccf_date>='" + rec_date_from + "' and sccf_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP1")

# TDP-02
rec_date_from = current_month + "-11"
rec_date_to = current_month + "-20"
tdp_df = duckdb.query("select * from sccf_df where sccf_date>='" + rec_date_from + "' and sccf_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP2")

# TDP-03
rec_date_from = current_month + "-21"
rec_date_to = sccf_dates[len(sccf_dates)-1]
tdp_df = duckdb.query("select * from sccf_df where sccf_date>='" + rec_date_from + "' and sccf_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP3")


# In[11]:


## trend ##

# path
path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/SCCF Matrices/SCCF_2by2_Matrix_' + rec_date_from[0:7] + '.xlsx'

# sheets
sheets = pd.ExcelFile(path).sheet_names

# trend data
qry = '''
select sccf_date, basepack, town, sum(inv_qty)*1.00/sum(ord_qty) sccf
from sccf_df
group by 1, 2, 3
union all
select sccf_date, 'overall' basepack, town, sum(inv_qty)*1.00/sum(ord_qty) sccf
from sccf_df
group by 1, 2, 3
union all
select sccf_date, basepack, 'overall' town, sum(inv_qty)*1.00/sum(ord_qty) sccf
from sccf_df
group by 1, 2, 3
'''
trend_df = duckdb.query(qry).df()
trend_df_piv = pd.pivot_table(trend_df.fillna(-0.01), values='sccf', index=['town', 'basepack'], columns='sccf_date', sort=False).fillna(-0.01)

# benchmark
qry = '''
select sum(inv_qty)*1.00/sum(ord_qty) benchmark_sccf
from sccf_df
'''
benchmark = duckdb.query(qry).df()['benchmark_sccf'].tolist()[0] 

# load
book = load_workbook(path)

# writer
writer = pd.ExcelWriter(path, engine = 'openpyxl')

# create, retrieve, remove sheets
sheet_name = rec_date_from[0:7] + '-trends'
writer.book = book
if sheet_name in sheets: del writer.book[sheet_name]

# write
trend_df_piv.reset_index().to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index = False)

# adjust
worksheet = writer.sheets[sheet_name]
it = 1
for column_cells in worksheet.columns:
    if it in [1, 2]: length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)  
    else: length = len(str(column_cells[0].value))
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    it = it + 1
writer.close()

# format
workbook = load_workbook(path)
worksheet = workbook[sheet_name]

# freeze
worksheet.freeze_panes = worksheet['C2']

# filter
worksheet.auto_filter.ref = worksheet.dimensions

# percent
column_letters = [col.column_letter for col in worksheet[1]]
for col in column_letters: 
    for cel in worksheet[(col)]: 
        cel.number_format = "0.00%" 
        
# color
font = Font(bold = True, color = 'EE1111')
dxf = DifferentialStyle(font = font)
rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, benchmark], dxf = dxf)
worksheet.conditional_formatting.add('C5:AH50000', rule)
    
# save
workbook.save(path)

# tabs in sheet
print("Worksheets in workbook: ")
print(pd.ExcelFile(path).sheet_names)


# In[12]:


# report
elapsed_time = str(round((time.time() - start_time) / 60.00, 2))
print("Elapsed time to run script (mins): " + elapsed_time)


# In[13]:


# analysis
qry = '''
select * 
from 
    (select * 
    from 
        (select 
            basepack, 
            town, 
            sccf_date, 
            lead(sccf_date, 1) over(partition by basepack, town order by sccf_date) sccf_date_1,
            lead(sccf_date, 2) over(partition by basepack, town order by sccf_date) sccf_date_2,
            lead(sccf_date, 3) over(partition by basepack, town order by sccf_date) sccf_date_3,
            lead(sccf_date, 4) over(partition by basepack, town order by sccf_date) sccf_date_4
        from (select basepack, town, sccf, sccf_date::date sccf_date from trend_df) tbl1
        where 
            sccf<0.90
            and town not in('WATER-DHAKA', 'MODERN TRADE', 'SMT & SHOPPING COMPL', 'OOH DISTRIBUTOR DHAK', 'OOH DISTRIBUTOR DHAKA')
        ) tbl1 
    where 
        sccf_date_4=(select max(sccf_date) from trend_df)
        and sccf_date_1-sccf_date=1
        and sccf_date_2-sccf_date_1=1
        and sccf_date_3-sccf_date_2=1
        and sccf_date_4-sccf_date_3=1
        and town='TEJGAON'
    ) tbl1 
    inner join 
    tgt_df tbl2 using(basepack, town)
order by tgt_cr desc
limit 5
'''
df = duckdb.query(qry).df()

# send - only me grp
import pywhatkit
emo = ":unlock\t"
msg = emo + " Are you continuously losing SCCF in UBL's very own Tejgaon? Here are packs, with < 90% SCCF for the past 5 days:\n- " + "\n- ".join(df['basepack'].tolist()) + "\nFind out more from this month's 2*2 SCCF matrix!"
print("\n" + msg)
pywhatkit.sendwhatmsg_to_group_instantly(group_id="DXqnN42tpV27ZoVWszBH9D", message=msg, tab_close=True)


# In[ ]:

