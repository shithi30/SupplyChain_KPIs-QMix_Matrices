#!/usr/bin/env python
# coding: utf-8

# In[1]:


# import
import pandas as pd
import duckdb
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import PatternFill, Font
import win32com.client
from win32com.client import Dispatch
from datetime import datetime
import time


# In[2]:


# read target/Stock Alloc
def read_tgt_file():
    file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/RPL Inputs/" + "February'24 Town x SKU Stock Allocation_National.xlsx"
    sheet_name = "Town x SKU x Case x TGT "
    df = pd.read_excel(open(file, "rb"), sheet_name=sheet_name, header=2, index_col=None)
    tgt_df = df[['CATEGORY', 'TOWN NAME', 'SKU NAME', 'ALLOCATON [CS]', 'TOWN x SKU TGT - TP Cr.']]
    tgt_df.columns = ['category', 'town', 'basepack', 'tgt_cs', 'tgt_cr']
    tgt_df = duckdb.query('''select upper(category) category, upper(town) town, upper(basepack) basepack, tgt_cs, tgt_cr from tgt_df''').df()
    return tgt_df


# In[3]:


# read achievement
def fetch_read_ach(rec_date_from, rec_date_to): 
    
    # inputs
    subject_pattern = 'Secondary CCFOT'

    # output folder
    output_dir = Path.cwd() / 'SCCF Inputs'
    output_dir.mkdir(parents=True, exist_ok=True)

    # output files
    filenames = set()

    # outlook inbox
    outlook = win32com.client.Dispatch('Outlook.Application').GetNamespace('MAPI')
    inbox = outlook.Folders.Item(1).Folders['Kader Bhai']

    # emails
    messages = inbox.Items
    for message in reversed(messages): 

        # subject
        subject = message.Subject
        if subject_pattern.lower() not in subject.lower(): continue

        # attachments
        attachments = message.Attachments
        for attachment in attachments:
            filename = attachment.FileName
            try: file_date = str(datetime.strptime(" ".join(filename.split()[3:6])[0:-9], '%d %b %Y'))[0:10]
            except: file_date = str(datetime.strptime(" ".join(filename.split()[3:6])[0:-9], '%d %B %Y'))[0:10]
            if file_date >= rec_date_from and file_date <= rec_date_to:
                filenames.add(filename)
                attachment.SaveAsFile(output_dir / filename)
                
    # read
    ach_df = pd.DataFrame()
    for f in filenames:
        print("Reading: " + f)
        file = "C:/Users/Shithi.Maitra/Unilever Codes/Ad Hoc/2by2 Matrices/SCCF Inputs/" + f
        df = pd.read_excel(open(file, "rb"), sheet_name="Sheet1", header=1, index_col=None)
        df = df[['Local Sales Region 4', 'Pack Size', 'CS.1']]
        df.columns = ['town', 'basepack', 'inv_qty']
        df = duckdb.query('''select upper(town) town, upper(basepack) basepack, inv_qty from df''').df()
        try: df['ach_date'] = str(datetime.strptime(" ".join(f.split()[3:6])[0:-9], '%d %B %Y'))[0:10]
        except: df['ach_date'] = str(datetime.strptime(" ".join(f.split()[3:6])[0:-9], '%d %b %Y'))[0:10]
        ach_df = ach_df.append(df)
    return ach_df


# In[4]:


# prepare report
def prepare_report(rec_date_from, rec_date_to, ach_df, sheet_name):
    
    # prepare achievement
    ach_df = ach_df.fillna(0)
    qry = '''
    select 
        case when basepack is not null then basepack else 'unspecified' end basepack, 
        case when category is not null then category else 'unspecified' end category, 
        case when town is not null then town else 'unspecified' end town, 
        inv_qty, tgt_cs,
        case when bp_tgt_cr*1.00/tot_tgt_cr is null then 0 else bp_tgt_cr*1.00/tot_tgt_cr end bp_bc, 
        case when town_tgt_cr*1.00/tot_tgt_cr is null then 0 else town_tgt_cr*1.00/tot_tgt_cr end town_bc,
        inv_qty*1.00/tgt_cs ach
    from 
        (select basepack, town, sum(inv_qty) inv_qty
        from ach_df
        group by 1, 2
        ) tbl0

        full join

        (select basepack, sum(tgt_cr) bp_tgt_cr
        from tgt_df
        group by 1
        ) tbl2 using(basepack)

        full join

        (select town, sum(tgt_cr) town_tgt_cr
        from tgt_df
        group by 1
        ) tbl3 using(town)

        left join 

        (select distinct basepack, category
        from tgt_df
        ) tbl6 using(basepack)
        
        left join

        (select basepack, town, sum(tgt_cs) tgt_cs
        from tgt_df
        group by 1, 2
        ) tbl4 using(basepack, town), 

        (select sum(tgt_cr) tot_tgt_cr
        from tgt_df
        ) tbl5
    order by bp_bc desc
    '''
    ach_df_contrib = duckdb.query(qry).df()

    # order columns
    qry = '''
    select 
        town, 
        max(town_bc) town_bc, 
        sum(max(town_bc)) over(order by max(town_bc) desc) town_bc_cum,
        sum(inv_qty)*1.00/sum(tgt_cs) town_ach
    from ach_df_contrib
    group by 1
    order by town_bc desc
    '''
    ord_df = duckdb.query(qry).df()
    town_ord = ord_df['town'].tolist()

    # town BC
    town_bc_cum_ord = ['town_bc_cum'] + ord_df['town_bc_cum'].tolist()
    town_bc_cum_ord = pd.DataFrame([town_bc_cum_ord])
    town_bc_ord = ['town_bc'] + ord_df['town_bc'].tolist()
    town_bc_ord = pd.DataFrame([town_bc_ord])
    town_ach_ord = ['town_ach'] + ord_df['town_ach'].tolist()
    town_ach_ord = pd.DataFrame([town_ach_ord])

    # basepack BC
    qry = '''
    select 
        category, 
        basepack, 
        max(bp_bc) basepack_bc, 
        sum(max(bp_bc)) over(order by max(bp_bc) desc) basepack_bc_cum,
        sum(inv_qty)*1.00/sum(tgt_cs) basepack_ach
    from ach_df_contrib
    group by 1, 2
    order by basepack_bc desc
    '''
    bp_bc_df = duckdb.query(qry).df()[['basepack_bc_cum', 'basepack_bc', 'basepack_ach', 'category']]
    
    # national
    qry = '''
    select * 
    from 
        (select sum(inv_qty)*1.00/sum(tgt_cs) national_ach
        from ach_df_contrib
        ) tbl1, 
        (select sum(inv_qty)*1.00/sum(tgt_cs) national_ach_excluding_mtsmtwt
        from ach_df_contrib
        where town not in('WATER-DHAKA', 'MODERN TRADE', 'SMT & SHOPPING COMPL', 'OOH DISTRIBUTOR DHAK')
        ) tbl2
    '''
    national_df = duckdb.query(qry).df()

    # pivot
    ach_df_piv = pd.pivot_table(ach_df_contrib.fillna(-0.01), values='ach', index=['basepack'], columns='town', sort=False).fillna(-0.01)
    ach_df_piv = ach_df_piv[town_ord]
    
    # path
    path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/Sec Ach Matrices/Sec_Ach_Matrix_' + rec_date_from[0:7] + '.xlsx'

    # if exists
    if_exists = 1
    try: book = load_workbook(path)
    except: if_exists = 0

    # writer
    writer = pd.ExcelWriter(path, engine = 'openpyxl')

    # create/retrieve/remove sheet(s)
    if if_exists == 1: 
        writer.book = book
        if sheet_name in book.sheetnames:
            if rec_date_from != rec_date_to:
                try: del writer.book[sheet_name] # sheet exists 
                except: pass                     # sheet does not exist
            else: 
                writer.close()
                return
    
    # write
    national_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=3, index=False)
    town_bc_cum_ord.to_excel(writer, sheet_name=sheet_name, startrow=2, startcol=4, header=False, index=False)
    town_bc_ord.to_excel(writer, sheet_name=sheet_name, startrow=3, startcol=4, header=False, index=False)
    town_ach_ord.to_excel(writer, sheet_name=sheet_name, startrow=4, startcol=4, header=False, index=False)
    ach_df_piv.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=4)
    bp_bc_df.to_excel(writer, sheet_name=sheet_name, startrow=5, startcol=0, index=False)

    # adjust
    worksheet = writer.sheets[sheet_name]
    for column_cells in worksheet.columns:
        length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)        
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    writer.close()
    
    # format
    workbook = load_workbook(path)
    worksheet = workbook[sheet_name]
    
    # percent
    column_letters = [col.column_letter for col in worksheet[1]]
    for col in column_letters: 
        for cel in worksheet[(col)]: 
            cel.number_format = "0.00%"
    
    # color
    font = Font(bold = True, color = 'EE1111')
    dxf = DifferentialStyle(font = font)
    rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, national_df['national_ach'].tolist()[0]], dxf = dxf)
    worksheet.conditional_formatting.add('C5:GZ300', rule)
    
    # freeze
    worksheet.freeze_panes = worksheet['F7']
    
    # save
    workbook.save(path)

    # tabs in sheet
    print("Worksheets in workbook: ")
    print(pd.ExcelFile(path).sheet_names)
    print()


# In[5]:


# record
start_time = time.time()


# In[6]:


# month, dates, target
tgt_df = read_tgt_file()

current_month = datetime.today().strftime('%Y-%m-%d')[0:7]
qry = '''
select (concat(left(current_date, 7), '-01')::date + generate_series::int)::text ach_date
from (select * from generate_series(0, 100)) tbl1 
where generate_series < date_part('day', current_date)
'''
ach_dates = duckdb.query(qry).df()['ach_date'].tolist()
print(ach_dates)


# In[7]:


# achievement
rec_date_from = ach_dates[0]
rec_date_to = ach_dates[len(ach_dates)-1]
ach_df = fetch_read_ach(rec_date_from, rec_date_to)


# In[8]:


# daily
for ach_date in ach_dates:
    day_df = duckdb.query("select * from ach_df where ach_date='" + ach_date + "'").df()
    if day_df.shape[0] == 0: continue
    prepare_report(ach_date, ach_date, day_df, ach_date)


# In[9]:


# MTD
prepare_report(rec_date_from, rec_date_to, ach_df, current_month + "-MTD")


# In[14]:


# TDP-01
rec_date_from = current_month + "-01"
rec_date_to = current_month + "-10"
tdp_df = duckdb.query("select * from ach_df where ach_date>='" + rec_date_from + "' and ach_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP1")

# TDP-02
rec_date_from = current_month + "-11"
rec_date_to = current_month + "-20"
tdp_df = duckdb.query("select * from ach_df where ach_date>='" + rec_date_from + "' and ach_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP2")

# TDP-03
rec_date_from = current_month + "-21"
rec_date_to = ach_dates[len(ach_dates)-1]
tdp_df = duckdb.query("select * from ach_df where ach_date>='" + rec_date_from + "' and ach_date<='" + rec_date_to + "'").df()
prepare_report(rec_date_from, rec_date_to, tdp_df, current_month + "-TDP3")


# In[11]:


## trend ##

# path
path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/Sec Ach Matrices/Sec_Ach_Matrix_' + current_month + '.xlsx'

# sheets
sheets = pd.ExcelFile(path).sheet_names

# trend data
qry = '''
with 
    tbl as
    (select *
    from 
        (select ach_date, basepack, town, sum(inv_qty) inv_qty
        from ach_df
        group by 1, 2, 3
        ) tbl1 

        inner join 

        (select basepack, town, sum(tgt_cs) tgt_cs
        from tgt_df
        group by 1, 2
        ) tbl2 using(basepack, town)
    ) 
select ach_date, basepack, town, inv_qty*1.00/tgt_cs ach from tbl
union all
select ach_date, 'overall' basepack, town, sum(inv_qty)*1.00/sum(tgt_cs) ach from tbl group by 1, 2, 3
union all
select ach_date, basepack, 'overall' town, sum(inv_qty)*1.00/sum(tgt_cs) ach from tbl group by 1, 2, 3
'''
trend_df = duckdb.query(qry).df()
trend_df_piv = pd.pivot_table(trend_df.fillna(-0.01), values='ach', index=['town', 'basepack'], columns='ach_date', sort=False).fillna(-0.01)

# benchmark
qry = '''
select inv_qty*1.00/tgt_cs benchmark_ach
from 
    (select sum(inv_qty) inv_qty from ach_df) tbl1, 
    (select sum(tgt_cs) tgt_cs from tgt_df) tbl2
'''
benchmark = duckdb.query(qry).df()['benchmark_ach'].tolist()[0] 

# load
book = load_workbook(path)

# writer
writer = pd.ExcelWriter(path, engine = 'openpyxl')

# create, retrieve, remove sheets
sheet_name = current_month + '-trends'
writer.book = book
if sheet_name in sheets: del writer.book[sheet_name]

# write
trend_df_piv.reset_index().to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index = False)

# adjust
worksheet = writer.sheets[sheet_name]
it = 1
for column_cells in worksheet.columns:
    if it in [1, 2]: length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)  
    else: length = len(str(column_cells[0].value))
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    it = it + 1
writer.close()

# format
workbook = load_workbook(path)
worksheet = workbook[sheet_name]

# freeze
worksheet.freeze_panes = worksheet['C2']

# filter
worksheet.auto_filter.ref = worksheet.dimensions

# percent
column_letters = [col.column_letter for col in worksheet[1]]
for col in column_letters: 
    for cel in worksheet[(col)]: 
        cel.number_format = "0.00%" 
        
# color
font = Font(bold = True, color = 'EE1111')
dxf = DifferentialStyle(font = font)
rule = Rule(type = 'cellIs', operator = 'between', formula = [0.001, benchmark], dxf = dxf)
worksheet.conditional_formatting.add('C5:AH50000', rule)
    
# save
workbook.save(path)

# tabs in sheet
print("Worksheets in workbook: ")
print(pd.ExcelFile(path).sheet_names)


# In[12]:


## raw ##

# path
path = 'C:/Users/Shithi.Maitra/OneDrive - Unilever/2d Matrices/Sec Ach Matrices/Sec_Ach_Matrix_' + current_month + '.xlsx'

# sheets
sheets = pd.ExcelFile(path).sheet_names

# raw data
qry = '''
select *
from 
    (select basepack, town, sum(inv_qty) inv_qty_till_date
    from ach_df
    group by 1, 2
    ) tbl1 

    full join 

    (select basepack, town, sum(tgt_cs) tgt_cs
    from tgt_df
    group by 1, 2
    ) tbl2 using(basepack, town)
'''
raw_df = duckdb.query(qry).df()

# load
book = load_workbook(path)

# writer
writer = pd.ExcelWriter(path, engine = 'openpyxl')

# create, retrieve, remove sheets
sheet_name = current_month + '-RAW'
writer.book = book
if sheet_name in sheets: del writer.book[sheet_name]

# write
raw_df.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index = False)

# adjust
worksheet = writer.sheets[sheet_name]
it = 1
for column_cells in worksheet.columns:
    if it in [1, 2]: length = max(len("".join(c for c in str(cell.value).replace(' ', 'A') if c.isalpha())) for cell in column_cells)  
    else: length = len(str(column_cells[0].value))
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length + 4
    it = it + 1
writer.close()

# tabs in sheet
print("Worksheets in workbook: ")
print(pd.ExcelFile(path).sheet_names)


# In[13]:


# report
elapsed_time = str(round((time.time() - start_time) / 60.00, 2))
print("Elapsed time to run script (mins): " + elapsed_time)


# In[ ]:

